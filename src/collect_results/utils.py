from src.utils import unzip_file, make_dirct, convert_bytes,select_key_with_value_condition, manual_flatten,float_to_percentage,assign_dict_key_empty_list
from src.CONSTANTS import benchmark_timeout
import os
import json
import glob
from shutil import copy
from tqdm import tqdm
import pandas as pd
from src.plots import scatter_plot,draw_cactus_plot_multiple_plotly
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
import shutil
green_fill = PatternFill(start_color="FF00FF00",
                           end_color="FF00FF00",
                           fill_type="solid")
import math


def summarize_excel_files():
    excel_files_dict = {
        "CEGAR-linear-union": ["uppmax-CEGAR-linear-fixed-heuristic-random",
                                 "uppmax-CEGAR-linear-union-linear-model-rank",
                                 "uppmax-CEGAR-linear-union-linear-model-score",
                                 "uppmax-CEGAR-linear-union-linear-model-SEHPlus",
                                 "uppmax-CEGAR-linear-union-linear-model-SEHMinus",
                                 "uppmax-CEGAR-linear-union-linear-model-REHPlus",
                                 "uppmax-CEGAR-linear-union-linear-model-REHMinus",
                                 "uppmax-CEGAR-linear-union-mixed-model-rank",
                                 "uppmax-CEGAR-linear-union-mixed-model-score",
                                 "uppmax-CEGAR-linear-union-mixed-model-SEHPlus",
                                 "uppmax-CEGAR-linear-union-mixed-model-SEHMinus",
                                 "uppmax-CEGAR-linear-union-mixed-model-REHPlus",
                                 "uppmax-CEGAR-linear-union-mixed-model-REHMinus",
                                 ],
        "symex-linear-union": ["uppmax-symex-linear-fixed-heuristic-random",
                                 "uppmax-symex-linear-union-linear-model-rank",
                                 "uppmax-symex-linear-union-linear-model-score",
                                 "uppmax-symex-linear-union-linear-model-SEHPlus",
                                 "uppmax-symex-linear-union-linear-model-SEHMinus",
                                 "uppmax-symex-linear-union-linear-model-REHPlus",
                                 "uppmax-symex-linear-union-linear-model-REHMinus",
                                 "uppmax-symex-linear-union-linear-model-twoQueue02",
                                 "uppmax-symex-linear-union-linear-model-twoQueue05",
                                 "uppmax-symex-linear-union-linear-model-twoQueue08",
                                 "uppmax-symex-linear-union-linear-model-schedule10",
                                 "uppmax-symex-linear-union-linear-model-schedule100",
                                 "uppmax-symex-linear-union-linear-model-schedule1000",
                                 "uppmax-symex-linear-union-mixed-model-rank",
                                 "uppmax-symex-linear-union-mixed-model-score",
                                 "uppmax-symex-linear-union-mixed-model-SEHPlus",
                                 "uppmax-symex-linear-union-mixed-model-SEHMinus",
                                 "uppmax-symex-linear-union-mixed-model-REHPlus",
                                 "uppmax-symex-linear-union-mixed-model-REHMinus",
                                 "uppmax-symex-linear-union-mixed-model-twoQueue02",
                                 "uppmax-symex-linear-union-mixed-model-twoQueue05",
                                 "uppmax-symex-linear-union-mixed-model-twoQueue08",
                                 "uppmax-symex-linear-union-mixed-model-schedule10",
                                 "uppmax-symex-linear-union-mixed-model-schedule100",
                                 "uppmax-symex-linear-union-mixed-model-schedule1000"
                                 ],
        "CEGAR-non-linear-union": ["uppmax-CEGAR-non-linear-fixed-heuristic-random",
                                   "uppmax-CEGAR-non-linear-union-non-linear-model-rank",
                                   "uppmax-CEGAR-non-linear-union-non-linear-model-score",
                                   "uppmax-CEGAR-non-linear-union-non-linear-model-SEHPlus",
                                   "uppmax-CEGAR-non-linear-union-non-linear-model-SEHMinus",
                                   "uppmax-CEGAR-non-linear-union-non-linear-model-REHPlus",
                                   "uppmax-CEGAR-non-linear-union-non-linear-model-REHMinus",
                                   "uppmax-CEGAR-non-linear-union-mixed-model-rank",
                                   "uppmax-CEGAR-non-linear-union-mixed-model-score",
                                   "uppmax-CEGAR-non-linear-union-mixed-model-SEHPlus",
                                   "uppmax-CEGAR-non-linear-union-mixed-model-SEHMinus",
                                   "uppmax-CEGAR-non-linear-union-mixed-model-REHPlus",
                                   "uppmax-CEGAR-non-linear-union-mixed-model-REHMinus",
                                   ],
        "symex-non-linear-union": ["uppmax-symex-non-linear-fixed-heuristic-random",
                                   "uppmax-symex-non-linear-union-non-linear-model-rank",
                                   "uppmax-symex-non-linear-union-non-linear-model-score",
                                   "uppmax-symex-non-linear-union-non-linear-model-SEHPlus",
                                   "uppmax-symex-non-linear-union-non-linear-model-SEHMinus",
                                   "uppmax-symex-non-linear-union-non-linear-model-REHPlus",
                                   "uppmax-symex-non-linear-union-non-linear-model-REHMinus",
                                   "uppmax-symex-non-linear-union-non-linear-model-twoQueue02",
                                   "uppmax-symex-non-linear-union-non-linear-model-twoQueue05",
                                   "uppmax-symex-non-linear-union-non-linear-model-twoQueue08",
                                   "uppmax-symex-non-linear-union-non-linear-model-schedule10",
                                   "uppmax-symex-non-linear-union-non-linear-model-schedule100",
                                   "uppmax-symex-non-linear-union-non-linear-model-schedule1000",
                                   "uppmax-symex-non-linear-union-mixed-model-rank",
                                   "uppmax-symex-non-linear-union-mixed-model-score",
                                   "uppmax-symex-non-linear-union-mixed-model-SEHPlus",
                                   "uppmax-symex-non-linear-union-mixed-model-SEHMinus",
                                   "uppmax-symex-non-linear-union-mixed-model-REHPlus",
                                   "uppmax-symex-non-linear-union-mixed-model-REHMinus",
                                   "uppmax-symex-non-linear-union-mixed-model-twoQueue02",
                                   "uppmax-symex-non-linear-union-mixed-model-twoQueue05",
                                   "uppmax-symex-non-linear-union-mixed-model-twoQueue08",
                                   "uppmax-symex-non-linear-union-mixed-model-schedule10",
                                   "uppmax-symex-non-linear-union-mixed-model-schedule100",
                                   "uppmax-symex-non-linear-union-mixed-model-schedule1000"
                                   ],

        "CEGAR-linear-minimal": ["uppmax-CEGAR-linear-fixed-heuristic-random",
                                 "uppmax-CEGAR-linear-minimal-linear-model-rank",
                                 "uppmax-CEGAR-linear-minimal-linear-model-score",
                                 "uppmax-CEGAR-linear-minimal-linear-model-SEHPlus", "uppmax-CEGAR-linear-minimal-linear-model-SEHMinus",
                                 "uppmax-CEGAR-linear-minimal-linear-model-REHPlus", "uppmax-CEGAR-linear-minimal-linear-model-REHMinus",
                                 "uppmax-CEGAR-linear-minimal-mixed-model-rank",
                                 "uppmax-CEGAR-linear-minimal-mixed-model-score",
                                 "uppmax-CEGAR-linear-minimal-mixed-model-SEHPlus",
                                 "uppmax-CEGAR-linear-minimal-mixed-model-SEHMinus",
                                 "uppmax-CEGAR-linear-minimal-mixed-model-REHPlus",
                                 "uppmax-CEGAR-linear-minimal-mixed-model-REHMinus",
                                 ],
        "symex-linear-minimal": ["uppmax-symex-linear-fixed-heuristic-random",
                                 "uppmax-symex-linear-minimal-linear-model-rank",
                                 "uppmax-symex-linear-minimal-linear-model-score",
                                 "uppmax-symex-linear-minimal-linear-model-SEHPlus", "uppmax-symex-linear-minimal-linear-model-SEHMinus",
                                 "uppmax-symex-linear-minimal-linear-model-REHPlus", "uppmax-symex-linear-minimal-linear-model-REHMinus",
                                 "uppmax-symex-linear-minimal-linear-model-twoQueue02",
                                 "uppmax-symex-linear-minimal-linear-model-twoQueue05",
                                 "uppmax-symex-linear-minimal-linear-model-twoQueue08",
                                 "uppmax-symex-linear-minimal-linear-model-schedule10",
                                 "uppmax-symex-linear-minimal-linear-model-schedule100",
                                 "uppmax-symex-linear-minimal-linear-model-schedule1000",
                                 "uppmax-symex-linear-minimal-mixed-model-rank",
                                 "uppmax-symex-linear-minimal-mixed-model-score",
                                 "uppmax-symex-linear-minimal-mixed-model-SEHPlus",
                                 "uppmax-symex-linear-minimal-mixed-model-SEHMinus",
                                 "uppmax-symex-linear-minimal-mixed-model-REHPlus",
                                 "uppmax-symex-linear-minimal-mixed-model-REHMinus",
                                 "uppmax-symex-linear-minimal-mixed-model-twoQueue02",
                                 "uppmax-symex-linear-minimal-mixed-model-twoQueue05",
                                 "uppmax-symex-linear-minimal-mixed-model-twoQueue08",
                                 "uppmax-symex-linear-minimal-mixed-model-schedule10",
                                 "uppmax-symex-linear-minimal-mixed-model-schedule100",
                                 "uppmax-symex-linear-minimal-mixed-model-schedule1000"
                                 ],
        "CEGAR-non-linear-minimal": ["uppmax-CEGAR-non-linear-fixed-heuristic-random",
                                     "uppmax-CEGAR-non-linear-minimal-non-linear-model-rank",
                                     "uppmax-CEGAR-non-linear-minimal-non-linear-model-score",
                                     "uppmax-CEGAR-non-linear-minimal-non-linear-model-SEHPlus", "uppmax-CEGAR-non-linear-minimal-non-linear-model-SEHMinus",
                                     "uppmax-CEGAR-non-linear-minimal-non-linear-model-REHPlus", "uppmax-CEGAR-non-linear-minimal-non-linear-model-REHMinus",
                                     "uppmax-CEGAR-non-linear-minimal-mixed-model-rank",
                                     "uppmax-CEGAR-non-linear-minimal-mixed-model-score",
                                     "uppmax-CEGAR-non-linear-minimal-mixed-model-SEHPlus",
                                     "uppmax-CEGAR-non-linear-minimal-mixed-model-SEHMinus",
                                     "uppmax-CEGAR-non-linear-minimal-mixed-model-REHPlus",
                                     "uppmax-CEGAR-non-linear-minimal-mixed-model-REHMinus",
                                     ],
        "symex-non-linear-minimal": ["uppmax-symex-non-linear-fixed-heuristic-random",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-rank",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-score",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-SEHPlus", "uppmax-symex-non-linear-minimal-non-linear-model-SEHMinus",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-REHPlus", "uppmax-symex-non-linear-minimal-non-linear-model-REHMinus",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-twoQueue02",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-twoQueue05",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-twoQueue08",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-schedule10",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-schedule100",
                                     "uppmax-symex-non-linear-minimal-non-linear-model-schedule1000",
                                     "uppmax-symex-non-linear-minimal-mixed-model-rank",
                                     "uppmax-symex-non-linear-minimal-mixed-model-score",
                                     "uppmax-symex-non-linear-minimal-mixed-model-SEHPlus",
                                     "uppmax-symex-non-linear-minimal-mixed-model-SEHMinus",
                                     "uppmax-symex-non-linear-minimal-mixed-model-REHPlus",
                                     "uppmax-symex-non-linear-minimal-mixed-model-REHMinus",
                                     "uppmax-symex-non-linear-minimal-mixed-model-twoQueue02",
                                     "uppmax-symex-non-linear-minimal-mixed-model-twoQueue05",
                                     "uppmax-symex-non-linear-minimal-mixed-model-twoQueue08",
                                     "uppmax-symex-non-linear-minimal-mixed-model-schedule10",
                                     "uppmax-symex-non-linear-minimal-mixed-model-schedule100",
                                     "uppmax-symex-non-linear-minimal-mixed-model-schedule1000"
                                     ],

        "CEGAR-linear-common": ["uppmax-CEGAR-linear-fixed-heuristic-random",
                                "uppmax-CEGAR-linear-common-linear-model-rank",
                                "uppmax-CEGAR-linear-common-linear-model-score",
                                 "uppmax-CEGAR-linear-common-linear-model-SEHPlus",
                                 "uppmax-CEGAR-linear-common-linear-model-SEHMinus",
                                 "uppmax-CEGAR-linear-common-linear-model-REHPlus",
                                 "uppmax-CEGAR-linear-common-linear-model-REHMinus",
                                 "uppmax-CEGAR-linear-common-mixed-model-rank",
                                 "uppmax-CEGAR-linear-common-mixed-model-score",
                                 "uppmax-CEGAR-linear-common-mixed-model-SEHPlus",
                                 "uppmax-CEGAR-linear-common-mixed-model-SEHMinus",
                                 "uppmax-CEGAR-linear-common-mixed-model-REHPlus",
                                 "uppmax-CEGAR-linear-common-mixed-model-REHMinus",
                                 ],
        "symex-linear-common": ["uppmax-symex-linear-fixed-heuristic-random",
                                "uppmax-symex-linear-common-linear-model-rank",
                                "uppmax-symex-linear-common-linear-model-score",
                                 "uppmax-symex-linear-common-linear-model-SEHPlus",
                                 "uppmax-symex-linear-common-linear-model-SEHMinus",
                                 "uppmax-symex-linear-common-linear-model-REHPlus",
                                 "uppmax-symex-linear-common-linear-model-REHMinus",
                                "uppmax-symex-linear-common-linear-model-schedule10",
                                "uppmax-symex-linear-common-linear-model-schedule100",
                                "uppmax-symex-linear-common-linear-model-schedule1000",
                                 "uppmax-symex-linear-common-mixed-model-rank",
                                 "uppmax-symex-linear-common-mixed-model-score",
                                 "uppmax-symex-linear-common-mixed-model-SEHPlus",
                                 "uppmax-symex-linear-common-mixed-model-SEHMinus",
                                 "uppmax-symex-linear-common-mixed-model-REHPlus",
                                 "uppmax-symex-linear-common-mixed-model-REHMinus",
                                 "uppmax-symex-linear-common-mixed-model-twoQueue02",
                                 "uppmax-symex-linear-common-mixed-model-twoQueue05",
                                 "uppmax-symex-linear-common-mixed-model-twoQueue08",
                                "uppmax-symex-linear-common-mixed-model-schedule10",
                                "uppmax-symex-linear-common-mixed-model-schedule100",
                                "uppmax-symex-linear-common-mixed-model-schedule1000"
                                 ],
        "CEGAR-non-linear-common": ["uppmax-CEGAR-non-linear-fixed-heuristic-random",
                                    "uppmax-CEGAR-non-linear-common-non-linear-model-rank",
                                    "uppmax-CEGAR-non-linear-common-non-linear-model-score",
                                     "uppmax-CEGAR-non-linear-common-non-linear-model-SEHPlus",
                                     "uppmax-CEGAR-non-linear-common-non-linear-model-SEHMinus",
                                     "uppmax-CEGAR-non-linear-common-non-linear-model-REHPlus",
                                     "uppmax-CEGAR-non-linear-common-non-linear-model-REHMinus",
                                     "uppmax-CEGAR-non-linear-common-mixed-model-rank",
                                     "uppmax-CEGAR-non-linear-common-mixed-model-score",
                                     "uppmax-CEGAR-non-linear-common-mixed-model-SEHPlus",
                                     "uppmax-CEGAR-non-linear-common-mixed-model-SEHMinus",
                                     "uppmax-CEGAR-non-linear-common-mixed-model-REHPlus",
                                     "uppmax-CEGAR-non-linear-common-mixed-model-REHMinus",
                                     ],
        "symex-non-linear-common": ["uppmax-symex-non-linear-fixed-heuristic-random",
                                    "uppmax-symex-non-linear-common-non-linear-model-rank",
                                    "uppmax-symex-non-linear-common-non-linear-model-score",
                                     "uppmax-symex-non-linear-common-non-linear-model-SEHPlus",
                                     "uppmax-symex-non-linear-common-non-linear-model-SEHMinus",
                                     "uppmax-symex-non-linear-common-non-linear-model-REHPlus",
                                     "uppmax-symex-non-linear-common-non-linear-model-REHMinus",
                                     "uppmax-symex-non-linear-common-non-linear-model-twoQueue02",
                                     "uppmax-symex-non-linear-common-non-linear-model-twoQueue05",
                                     "uppmax-symex-non-linear-common-non-linear-model-twoQueue08",
                                    "uppmax-symex-non-linear-common-non-linear-model-schedule10",
                                    "uppmax-symex-non-linear-common-non-linear-model-schedule100",
                                    "uppmax-symex-non-linear-common-non-linear-model-schedule1000",
                                     "uppmax-symex-non-linear-common-mixed-model-rank",
                                     "uppmax-symex-non-linear-common-mixed-model-score",
                                     "uppmax-symex-non-linear-common-mixed-model-SEHPlus",
                                     "uppmax-symex-non-linear-common-mixed-model-SEHMinus",
                                     "uppmax-symex-non-linear-common-mixed-model-REHPlus",
                                     "uppmax-symex-non-linear-common-mixed-model-REHMinus",
                                     "uppmax-symex-non-linear-common-mixed-model-twoQueue02",
                                     "uppmax-symex-non-linear-common-mixed-model-twoQueue05",
                                     "uppmax-symex-non-linear-common-mixed-model-twoQueue08",
                                     "uppmax-symex-non-linear-common-mixed-model-schedule10",
                                     "uppmax-symex-non-linear-common-mixed-model-schedule100",
                                     "uppmax-symex-non-linear-common-mixed-model-schedule1000"
                                     ],

        # "CEGAR-linear-train+valid-union": ["uppmax-CEGAR-linear-train+valid-union-random-869",
        #                                    "uppmax-CEGAR-linear-train+valid-union-label-869"],
        # "symex-linear-train+valid-union": ["uppmax-symex-linear-train+valid-union-random-869",
        #                                    "uppmax-symex-linear-train+valid-union-label-869"],
        # "CEGAR-linear-train+valid-minimal": ["uppmax-CEGAR-linear-train+valid-minimal-random-861",
        #                                      "uppmax-CEGAR-linear-train+valid-minimal-label-861"],
        # "symex-linear-train+valid-minimal": ["uppmax-symex-linear-train+valid-minimal-random-861",
        #                                      "uppmax-symex-linear-train+valid-minimal-label-861"],
        # "CEGAR-non-linear-train+valid-union": ["uppmax-CEGAR-non-linear-train+valid-union-label-1797",
        #                                        "uppmax-CEGAR-non-linear-train+valid-union-random-1797"],
        # "symex-non-linear-train+valid-union": ["uppmax-symex-non-linear-train+valid-union-random-1797",
        #                                        "uppmax-symex-non-linear-train+valid-union-label-1797"],
    }

    # non-linear
    # excel_files = ["uppmax-CEGAR-non-linear-train+valid-union-label-1797"]# CEGAR train+valid
    # excel_files = ["uppmax-synex-non-linear-train+valid-union-random-1797","uppmax-symex-non-linear-train+valid-union-label-1797"]  # symex train+valid

    summary_file = "/home/cheli243/PycharmProjects/HintsLearning/benchmarks/final-linear-evaluation/data_summary/summary.xlsx"
    #write summary excel
    # with pd.ExcelWriter(summary_file) as writer:
    #     pd.DataFrame(pd.DataFrame({})).to_excel(writer, sheet_name="summary")
    #
    #     for k in excel_files_dict:
    #         excel_files = excel_files_dict[k]
    #         columns = ["category"] + ["total","original_solved","original_safe", "original_unsafe", "original_avg_t","original_avg_t_s","original_avg_t_cs","original_avg_t_ocs","original_avg_t_safe","original_avg_t_unsafe"] + manual_flatten(
    #             [[f + "_solved",f + "_safe", f + "_unsafe", f + "_avg_t",f + "_avg_t_s",f + "_avg_t_cs",f+"_avg_t_ocs",f+"_avg_t_safe",f+"_avg_t_unsafe"] for f in excel_files])
    #         output_dict = {x: [] for x in columns}
    #         engine = "symex" if "symex" in excel_files[0] else "CEGAR"
    #
    #         # get original safe and unsafe
    #         solvability_dict = read_solvability_dict(
    #             "/home/cheli243/PycharmProjects/HintsLearning/benchmarks/final-linear-evaluation/data_summary/" +
    #             excel_files[0] + ".xlsx",
    #             sheet_name="category_summary")
    #
    #         output_dict["total"] = ["total"] + solvability_dict["number_predicted"]
    #
    #         output_dict["original_solved"] = ["solved"] + [a + b for a, b in
    #                                                      zip(solvability_dict["eldarica_" + engine + "_original_safe"],
    #                                                          solvability_dict[
    #                                                              "eldarica_" + engine + "_original_unsafe"])]
    #         output_dict["original_safe"] = ["safe"] + solvability_dict["eldarica_" + engine + "_original_safe"]
    #         output_dict["original_unsafe"] = ["unsafe"] + solvability_dict["eldarica_" + engine + "_original_unsafe"]
    #         solving_time_list=solvability_dict["eldarica_" + engine + "_original_solving_time"]
    #         common_solving_time_list = solvability_dict["eldarica_" + engine + "_original_common_solving_time"]
    #         common_original_solving_time_list = solvability_dict["eldarica_" + engine + "_original_common_original_solving_time"]
    #         common_solving_count_list = solvability_dict["eldarica_" + engine + "_original_common_solving_count"]
    #         sat_solving_time_list=solvability_dict["eldarica_" + engine + "_original_sat_solving_time"]
    #         sat_solving_count_list=solvability_dict["eldarica_" + engine + "_original_safe"]
    #         unsat_solving_time_list=solvability_dict["eldarica_" + engine + "_original_unsat_solving_time"]
    #         unsat_solving_count_list=solvability_dict["eldarica_" + engine + "_original_unsafe"]
    #         average_solving_time_list,average_solving_time_solved_list,average_solving_common_solving_time_list,average_solving_common_original_solving_time_list,average_safe_solving_time_list,average_unsafe_solving_time_list = compute_average_solving_time(
    #             output_dict["total"],output_dict["original_solved"] ,solving_time_list,common_solving_time_list,common_solving_count_list,common_original_solving_time_list,sat_solving_time_list,sat_solving_count_list,unsat_solving_time_list,unsat_solving_count_list)
    #         output_dict["original_avg_t"] = ["avg_t"] + [x if isinstance(x,str) else format(x,".2f") for x in average_solving_time_list]
    #         output_dict["original_avg_t_s"] = ["avg_t_s"] + [x if isinstance(x,str) else format(x,".2f") for x in average_solving_time_solved_list]
    #         output_dict["original_avg_t_cs"] = ["avg_t_cs"] + [x if isinstance(x, str) else format(x, ".2f") for x in average_solving_common_solving_time_list]
    #         output_dict["original_avg_t_ocs"] = ["avg_t_ocs"] + [x if isinstance(x, str) else format(x, ".2f") for x in average_solving_common_original_solving_time_list]
    #         output_dict["original_avg_t_safe"] = ["avg_t_safe"] + [x if isinstance(x, str) else format(x, ".2f") for x in average_safe_solving_time_list]
    #         output_dict["original_avg_t_unsafe"] = ["avg_t_unsafe"] + [x if isinstance(x, str) else format(x, ".2f") for x in average_unsafe_solving_time_list]
    #         output_dict["category"] = [" "] + solvability_dict["category"]
    #
    #
    #         # get safe and unsafe from other excels
    #         for f in excel_files:
    #             solvability_dict = read_solvability_dict(
    #                 "/home/cheli243/PycharmProjects/HintsLearning/benchmarks/final-linear-evaluation/data_summary/" + f + ".xlsx",
    #                 sheet_name="category_summary")
    #             output_dict[f + "_solved"] = ["solved"] + [a + b for a, b in zip(
    #                 solvability_dict["vb_eldarica_" + engine + "_prioritize_safe"],
    #                 solvability_dict["vb_eldarica_" + engine + "_prioritize_unsafe"])]
    #             output_dict[f + "_safe"] = ["safe"] + solvability_dict["vb_eldarica_" + engine + "_prioritize_safe"]
    #             output_dict[f + "_unsafe"] = ["unsafe"] + solvability_dict[
    #                 "vb_eldarica_" + engine + "_prioritize_unsafe"]
    #             solving_time_list=solvability_dict["vb_eldarica_" + engine + "_prioritize_solving_time"]
    #             common_solving_time_list = solvability_dict["vb_eldarica_" + engine + "_prioritize_common_solving_time"]
    #             common_original_solving_time_list = solvability_dict["vb_eldarica_" + engine + "_prioritize_common_original_solving_time"]
    #             common_solving_count_list = solvability_dict["vb_eldarica_" + engine + "_prioritize_common_solving_count"]
    #             sat_solving_time_list = solvability_dict["vb_eldarica_" + engine + "_prioritize_sat_solving_time"]
    #             sat_solving_count_list = solvability_dict["vb_eldarica_" + engine + "_prioritize_safe"]
    #             unsat_solving_time_list = solvability_dict["vb_eldarica_" + engine + "_prioritize_unsat_solving_time"]
    #             unsat_solving_count_list = solvability_dict["vb_eldarica_" + engine + "_prioritize_unsafe"]
    #             average_solving_time_list,average_solving_time_solved_list,average_solving_common_solving_time_list,average_solving_common_original_solving_time_list,average_safe_solving_time_list,average_unsafe_solving_time_list = compute_average_solving_time(output_dict["total"],output_dict[f + "_solved"]
    #             ,solving_time_list,common_solving_time_list,common_solving_count_list,common_original_solving_time_list,sat_solving_time_list,sat_solving_count_list,unsat_solving_time_list,unsat_solving_count_list)
    #             output_dict[f + "_avg_t"] = ["avg_t"] + [x if isinstance(x,str) else format(x,".2f") for x in average_solving_time_list]
    #             output_dict[f + "_avg_t_s"] = ["avg_t_s"] + [x if isinstance(x,str) else format(x,".2f") for x in average_solving_time_solved_list]
    #             output_dict[f + "_avg_t_cs"] = ["avg_t_cs"] + [x if isinstance(x, str) else format(x, ".2f") for x in average_solving_common_solving_time_list]
    #             output_dict[f + "_avg_t_ocs"] = ["avg_t_ocs"] + [x if isinstance(x, str) else format(x, ".2f") for x in average_solving_common_original_solving_time_list]
    #             output_dict[f + "_avg_t_safe"] = ["avg_t_safe"] + [x if isinstance(x, str) else format(x, ".2f") for x in average_safe_solving_time_list]
    #             output_dict[f + "_avg_t_unsafe"] = ["avg_t_unsafe"] + [x if isinstance(x, str) else format(x, ".2f") for x in average_unsafe_solving_time_list]
    #
    #         # for x in output_dict:
    #         #     print(x,len(output_dict[x]))
    #         pd.DataFrame(pd.DataFrame(output_dict)).to_excel(writer, sheet_name=k)
    #
    #
    # # merge some cells
    # for e_k in excel_files_dict:
    #     excel_files = excel_files_dict[e_k]
    #     # Load the Excel file
    #     workbook = load_workbook(summary_file)
    #     # Select the desired sheet
    #     sheet = workbook[e_k]
    #
    #     # Merge cells
    #     sheet.merge_cells('D1:L1')  # Merge cells in the range
    #     sheet["D1"].value = "Original"
    #
    #     merge_dict = {f: [] for f in excel_files}
    #     last_column_letter = sheet.dimensions.split(':')[1].strip('1234567890')
    #     for f in excel_files:
    #         for row in sheet["E1:" + last_column_letter + "1"]:
    #             for cell in row:
    #                 if (f + "_solved" == cell.value or f + "_safe" == cell.value or f + "_unsafe" == cell.value or f + "_avg_t" == cell.value or
    #                         f + "_avg_t_s" == cell.value or f + "_avg_t_cs" == cell.value or f + "_avg_t_ocs" == cell.value or f + "_avg_t_safe" == cell.value or f + "_avg_t_unsafe" == cell.value):
    #                     merge_dict[f].append(cell.coordinate)
    #     for k in merge_dict:
    #         sheet.merge_cells(merge_dict[k][0] + ":" + merge_dict[k][-1])
    #         sheet[merge_dict[k][0]].value = k.replace(e_k + "-", "").replace("uppmax-", "")
    #
    #     # add scatter plot inside
    #     count = 20 #row number
    #     for f in excel_files:
    #         img = Image(
    #             "/home/cheli243/PycharmProjects/HintsLearning/benchmarks/final-linear-evaluation/data_summary/" + f + ".png")
    #         sheet["A" + str(count)] = ''
    #         sheet["A" + str(count)].comment = None
    #         sheet.add_image(img, 'A' + str(count))
    #         count += 35 #row number
    #
    #     #compute improve percentage for solving time
    #     total_row = 13 if "non-linear" in e_k else 15
    #     column_number = 9
    #     sheet["B" + str(total_row + 2)].value = "improve percentage"
    #
    #     oirginal_column_number = 4  # solved
    #     compute_improved_percentage(sheet, oirginal_column_number, total_row, column_number)
    #     oirginal_column_number = 5  # safe
    #     compute_improved_percentage(sheet, oirginal_column_number, total_row, column_number)
    #     oirginal_column_number = 6  # unsafe
    #     compute_improved_percentage(sheet, oirginal_column_number, total_row, column_number)
    #     oirginal_st_column_number=7 #avg_t
    #     compute_improved_percentage_solving_time(sheet, oirginal_st_column_number, total_row, column_number)
    #     oirginal_st_column_number=8 #avg_t_s
    #     compute_improved_percentage_solving_time(sheet, oirginal_st_column_number, total_row, column_number)
    #     oirginal_st_column_number=9 #avg_t_cs
    #     compute_improved_percentage_for_common_solving_time(sheet, oirginal_st_column_number, total_row, column_number)
    #     oirginal_st_column_number = 11  # avg_t_safe
    #     compute_improved_percentage_solving_time(sheet, oirginal_st_column_number, total_row, column_number)
    #     oirginal_st_column_number = 12  # avg_t_unsafe
    #     compute_improved_percentage_solving_time(sheet, oirginal_st_column_number, total_row, column_number)
    #
    #     summary_max_for_each_sheet(sheet, total_row, column_number)
    #
    #     # Save the modified workbook
    #     workbook.save(summary_file)
    #
    # #write final summary, best strategy
    # workbook = load_workbook(summary_file)
    # sheet = workbook["summary"]
    # sheet.delete_rows(1, sheet.max_row) #clear the sheet
    # measurement_row_map = {"solved": 23,"safe":24,"unsafe":25,"avt_t":26,"avg_t_s":27,"avg_t_cs":28,"avg_t_safe":29,"avg_t_unsafe":30}
    # ce_types=["union", "minimal", "common"]
    # engine_data_list=[]
    # for e in ["CEGAR","symex"]:
    #     for d in ["linear","non-linear"]:
    #         engine_data_list.append(e + "-" + d)
    #
    # for engine_data_index,engine_data in enumerate(engine_data_list):
    #     for i,measurement in enumerate(measurement_row_map):
    #         write_one_block_summary_best_strategy(workbook, ce_types, measurement_row_map, measurement,i,engine_data,engine_data_index)
    #
    # # write final summary, best data set
    # column_number = 9
    # starting_row = 32
    # measurement_column_map = {"solved": 0, "safe": 1, "unsafe": 2, "avt_t": 3, "avg_t_s": 4, "avg_t_cs": 5,
    #                        "avg_t_safe": 7, "avg_t_unsafe": 8}
    # strategy_list_symex=["twoQueue02","twoQueue05","twoQueue08","schedule10","schedule100","schedule1000"]
    # strategy_list=["random","rank","score","SEHPlus","SEHMinus","REHPlus","REHMinus"]+strategy_list_symex
    #
    # #write measurement column
    # for index,measurement in enumerate(measurement_column_map):
    #     sheet = workbook["summary"]
    #     s_row = starting_row + len(strategy_list) * index
    #     e_row = starting_row + len(strategy_list) * (index+1) - 1
    #     sheet["A"+ str(s_row)].value = measurement
    #     sheet.merge_cells("A"+ str(s_row) + ':'+"A"+ str(e_row))
    # #write content
    # for engine_data_index, engine_data in enumerate(engine_data_list):
    #     total_row = 13 if "non-linear" in engine_data else 15
    #     for measurement in measurement_column_map:
    #         index = measurement_column_map[measurement]
    #         write_one_block_summary_best_data_set(workbook, ce_types, total_row, column_number, starting_row,
    #                                               strategy_list, measurement, index, engine_data,
    #                                               engine_data_index,strategy_list_symex,summary_column_number=5)
    #
    # workbook.save(summary_file)
    #
    # #cactus plots
    # for dataset in ["linear","non-linear"]:
    #     draw_cactus_plot_for_selected_configurations(dataset)

    # todo: write profolio
    for ce_type in ["union","minimal","common"]:
        collect_profolio(summary_file,ce_type)




def collect_profolio(summary_file,ce_type):
    folder="/home/cheli243/PycharmProjects/HintsLearning/benchmarks/final-linear-evaluation/data_summary/"
    strategy_list = ["rank", "score", "SEHPlus", "SEHMinus", "REHPlus", "REHMinus"]
    strategy_list_symex=strategy_list+["twoQueue02","twoQueue05","twoQueue08"]
    #read each block
    all_data_list=[]
    original_data_list=[]
    for engine in ["CEGAR","symex"]:
        for dataset in ["linear","non-linear"]:
            if engine=="CEGAR":
                strategies=strategy_list
            else:
                strategies=strategy_list_symex
            one_block_data_list=[]
            for strategy in strategies:
                one_configuration_file=folder+"uppmax-"+engine+"-"+dataset+"-"+ce_type+"-"+dataset+"-model-"+strategy+".xlsx"
                workbook = load_workbook(one_configuration_file)
                sheet = workbook["data"]
                one_block_data_list.append([[cell.value for cell in sheet["B"]][1:],[cell.value for cell in sheet["T"]][1:],[cell.value for cell in sheet["U"]][1:]])
            all_data_list.append(one_block_data_list)

            #read original for each block
            one_configuration_file = folder + "uppmax-"+engine+"-"+dataset+"-fixed-heuristic-random" + ".xlsx"
            workbook = load_workbook(one_configuration_file)
            sheet = workbook["data"]
            one_block_original_data_list=[]
            for n,sa,t in zip([cell.value for cell in sheet["B"]][1:], [cell.value for cell in sheet["R"]][1:],[cell.value for cell in sheet["S"]][1:]):
                one_block_original_data_list.append([n,sa,t])
            original_data_list.append(one_block_original_data_list)


    #compute profolio for each block
    workbook = load_workbook(summary_file)
    sheet = workbook["summary"]
    summary_column = 5
    for block_index,(one_block,one_block_original) in enumerate(zip(all_data_list,original_data_list)):

        one_block_profolio=[]
        for i, x in enumerate(one_block[0][0]):
            one_row_best_solving_time=100000
            one_row_best_satisfiability="None"
            one_row_best_name="None"
            for s in one_block:
                if s[2][i]<one_row_best_solving_time:
                    one_row_best_solving_time=s[2][i]
                    one_row_best_satisfiability=s[1][i]
                    one_row_best_name=s[0][i]
            one_block_profolio.append([one_row_best_name, one_row_best_satisfiability, one_row_best_solving_time])

        #compute columns
        safe_solving_time_list=[]
        unsafe_solving_time_list=[]
        unknown_list=[]
        original_safe_solving_time_list=[]
        original_unsafe_solving_time_list=[]
        original_unknown_list=[]
        common_solving_time_profolio_list=[]
        common_solving_time_original_list = []
        for row,row_original in zip(one_block_profolio,one_block_original):
            if row[1]=="safe":
                safe_solving_time_list.append(row[2])
            elif row[1]=="unsafe":
                unsafe_solving_time_list.append(row[2])
            else:
                unknown_list.append(row[2])
            if row_original[1]=="safe":
                original_safe_solving_time_list.append(row_original[2])
            elif row_original[1]=="unsafe":
                original_unsafe_solving_time_list.append(row_original[2])
            else:
                original_unknown_list.append(row_original[2])
            #compute common solved time
            if row[1]==row_original[1] and float(row[2])<benchmark_timeout:
                common_solving_time_profolio_list.append(row[2])
                common_solving_time_original_list.append(row_original[2])


        safe_count=len(safe_solving_time_list)
        unsafe_count=len(unsafe_solving_time_list)
        total_solved=safe_count+unsafe_count
        average_safe_solving_time=sum(safe_solving_time_list)/len(safe_solving_time_list)
        average_unsafe_solving_time=sum(unsafe_solving_time_list)/len(unsafe_solving_time_list)
        average_solving_time=sum(safe_solving_time_list+unsafe_solving_time_list+unknown_list)/len(safe_solving_time_list+unsafe_solving_time_list+unknown_list)
        average_common_solving_time=sum(common_solving_time_profolio_list)/len(common_solving_time_profolio_list)

        original_safe_count = len(original_safe_solving_time_list)
        original_unsafe_count = len(original_unsafe_solving_time_list)
        original_total_solved = original_safe_count + original_unsafe_count
        original_average_safe_solving_time = sum(original_safe_solving_time_list) / len(original_safe_solving_time_list)
        original_average_unsafe_solving_time = sum(original_unsafe_solving_time_list) / len(original_unsafe_solving_time_list)
        original_average_solving_time = sum(original_safe_solving_time_list + original_unsafe_solving_time_list + original_unknown_list) / len(original_safe_solving_time_list + original_unsafe_solving_time_list + original_unknown_list)
        average_common_solving_time_original = sum(common_solving_time_original_list) / len(common_solving_time_original_list)

        improved_percentage_safe_count=float_to_percentage((safe_count-original_safe_count)/original_safe_count)
        improved_percentage_unsafe_count=float_to_percentage((unsafe_count-original_unsafe_count)/original_unsafe_count)
        improved_percentage_total_solved=float_to_percentage((total_solved-original_total_solved)/original_total_solved)
        improved_percentage_average_safe_solving_time=float_to_percentage((original_average_safe_solving_time-average_safe_solving_time)/original_average_safe_solving_time)
        improved_percentage_average_unsafe_solving_time=float_to_percentage((original_average_unsafe_solving_time-average_unsafe_solving_time)/original_average_unsafe_solving_time)
        improved_percentage_average_solving_time=float_to_percentage((original_average_solving_time-average_solving_time)/original_average_solving_time)
        improved_percentage_average_common_solving_time=float_to_percentage(100*(average_common_solving_time_original-average_common_solving_time)/average_common_solving_time_original)

        dataset_row_start_dict={"union":3,"minimal":4,"common":5}
        dataset_row_start=dataset_row_start_dict[ce_type]
        #write back to excel
        strategy_row_content_dict={"solved":[dataset_row_start,total_solved,improved_percentage_total_solved],"safe":[dataset_row_start+3,safe_count,improved_percentage_safe_count],
                                   "unsafe":[dataset_row_start+6,unsafe_count,improved_percentage_unsafe_count],"avt_t":[dataset_row_start+9,average_solving_time,improved_percentage_average_solving_time],
                                   "avg_t_cs":[dataset_row_start+15,average_common_solving_time,improved_percentage_average_common_solving_time],
                                   "avg_t_safe":[dataset_row_start+18,average_safe_solving_time,improved_percentage_average_safe_solving_time],
                                   "avg_t_unsafe":[dataset_row_start+21,average_unsafe_solving_time,improved_percentage_average_unsafe_solving_time]}
        sheet[get_column_letter(summary_column*(block_index+1)  )+"2"].value="pf"
        sheet[get_column_letter(summary_column * (block_index + 1) + 1) + "2"].value = "pf_ip"
        for c in strategy_row_content_dict:
            sheet[get_column_letter(summary_column*(block_index+1))+str(strategy_row_content_dict[c][0])].value=strategy_row_content_dict[c][1]
            sheet[get_column_letter(summary_column * (block_index + 1)+1) + str(strategy_row_content_dict[c][0])].value =strategy_row_content_dict[c][2]


    workbook.save(summary_file)







def draw_cactus_plot_for_selected_configurations(dataset):
    summary_folder = "/home/cheli243/PycharmProjects/HintsLearning/benchmarks/final-linear-evaluation/data_summary"
    #linear cactus plot, union
    comparison_options_CEGAR=["REHPlus","REHMinus","SEHPlus","SEHMinus"]
    comparison_options_symex = comparison_options_CEGAR+["twoQueue08"]
    ledgends=[]
    lines=[]
    for engine in ["CEGAR","symex"]:
        # original+random
        excel_file=os.path.join(summary_folder,"uppmax-"+engine+"-"+dataset+"-fixed-heuristic-random.xlsx")
        workbook = load_workbook(excel_file)
        sheet = workbook["data"]
        ledgends.append(engine+"-"+"fixed")
        lines.append([cell.value for cell in sheet["S"]][1:])
        ledgends.append(engine+"-"+"random")
        lines.append([cell.value for cell in sheet["U"]][1:])
        # priority function options
        if engine =="CEGAR":
            for option in comparison_options_CEGAR:
                excel_file= os.path.join(summary_folder,"uppmax-"+engine+"-"+dataset+"-union-"+dataset+"-model-"+option+".xlsx")
                workbook = load_workbook(excel_file)
                sheet = workbook["data"]
                ledgends.append(engine + "-" + option)
                lines.append([cell.value for cell in sheet["U"]][1:])
        else:
            for option in comparison_options_symex:
                excel_file = os.path.join(summary_folder,"uppmax-" + engine + "-"+dataset+"-union-"+dataset+"-model-" + option + ".xlsx")
                workbook = load_workbook(excel_file)
                sheet = workbook["data"]
                ledgends.append(engine + "-" + option)
                lines.append([cell.value for cell in sheet["U"]][1:])

    plot_name="1-cactus-"+dataset+"-union"
    scale="log"
    x_axis_right_limit=240 if dataset=="linear" else 450
    draw_cactus_plot_multiple_plotly(summary_folder,plot_name,scale,lines, ledgends,x_axis_right_limit)

def write_one_block_summary_best_data_set(workbook,ce_types,total_row,column_number,starting_row,strategy_list,measurement,
                                          index,engine_data,engine_data_index,strategy_list_symex,summary_column_number):
    strategy_dict={}
    strategy_list=assign_dict_key_empty_list(strategy_dict,strategy_list)
    improved_percentage_row=total_row+2
    strategies={ss:i*column_number+4+index for i,ss in enumerate(strategy_dict)}
    strategies = {k:strategies[k]+column_number for k in strategies}
    max_percentage_value_list=[]
    #write numerical values
    for s in strategies:
        current_column_number = strategies[s]
        max_percentage_value=-1
        max_CE=""
        if "CEGAR" in engine_data and s in strategy_list_symex:
            max_percentage_value = -1
            max_CE = ""
        else:
            for ce in ce_types:
                sheet = workbook[engine_data+"-" + ce]
                improved_percentage_value = percentage_to_float(sheet[get_column_letter(current_column_number) + str(improved_percentage_row)].value)
                if improved_percentage_value > max_percentage_value:
                    max_percentage_value = improved_percentage_value
                    max_CE=ce
        strategy_dict[s].append([max_CE,float_to_percentage(max_percentage_value)])
        max_percentage_value_list.append(max_percentage_value)



    # write to summary sheet
    sheet = workbook["summary"]
    if measurement in ["avg_t_safe","avg_t_unsafe"]:
        index-=1
    s_row = starting_row + len(strategy_dict) * index

    for i,s in enumerate(strategy_dict):
        if "CEGAR" in engine_data and s in strategy_list_symex:
            sheet[get_column_letter((engine_data_index) * summary_column_number + 2) + str(i + s_row)].value = ""
            sheet[get_column_letter((engine_data_index) * summary_column_number + 3) + str(i + s_row)].value = ""
            sheet[get_column_letter((engine_data_index) * summary_column_number + 4) + str(i + s_row)].value = ""
        else:
            sheet[get_column_letter((engine_data_index) * summary_column_number + 2) + str(i + s_row)].value = s
            sheet[get_column_letter((engine_data_index)*summary_column_number+3) + str(i + s_row)].value = strategy_dict[s][0][0]
            sheet[get_column_letter((engine_data_index)*summary_column_number+4) +str(i+s_row)].value = strategy_dict[s][0][1]
        if max(max_percentage_value_list)==percentage_to_float(strategy_dict[s][0][1]):
            sheet[get_column_letter((engine_data_index)*summary_column_number+4) + str(i + s_row)].fill = green_fill



def write_one_block_summary_best_strategy(workbook,ce_types,measurement_row_map,measurement,measurement_index,engine_data,column_index):
    ce_number=3
    column_number=5
    strategy_list = []
    percentage_list = []
    for ce_type in ce_types:
        sheet = workbook[engine_data+ "-" + ce_type]
        strategy = sheet["R" + str(measurement_row_map[measurement])].value
        percentage = sheet["S" + str(measurement_row_map[measurement])].value
        strategy_list.append(strategy)
        percentage_list.append(percentage)

    #write to summary sheet
    sheet = workbook["summary"]
    #merge rows
    sheet["A"+str(measurement_index*ce_number+3)].value = measurement #solved, safe, unsafe, average time ...
    sheet.merge_cells('A'+str(measurement_index*ce_number+3)+':A'+str(measurement_index*ce_number+5))
    #merge columns
    sheet[get_column_letter(column_index*column_number+2)+"1"].value = engine_data #CEGAR-linear, CEGAR-non-linear, symex-linear...
    sheet.merge_cells(str(get_column_letter(column_index*column_number+2))+'1:'+str(get_column_letter(column_index*column_number+column_number+1))+'1')
    for i, (s, p, ce) in enumerate(zip(strategy_list, percentage_list, ce_types)):
        sheet[str(get_column_letter(column_index*column_number+2)) + str(i+1+measurement_index*ce_number + 2)].value = ce
        if "random" in s:
            s="random"
        sheet[str(get_column_letter(column_index*column_number+3)) + str(i+1+measurement_index*ce_number + 2)].value = s
        sheet[str(get_column_letter(column_index*column_number+4)) + str(i+1+measurement_index*ce_number + 2)].value = p
        if p == max(percentage_list):
            sheet[str(get_column_letter(column_index * column_number + 4)) + str(i+1 + measurement_index * ce_number + 2)].fill=green_fill

def summary_max_for_each_sheet(sheet,total_row,column_number):
    # color max value for improve percentage
    # Define the fill color (RGB)

    row_number_map = {"strategy": 1}
    column_number_map = {"solved": 13, "safe": 14, "unsafe": 15,"avg_t":16,"avg_t_s":17,"avg_t_cs":18,"avg_t_safe":20,"avg_t_unsafe":21}#

    summary_row = 22
    max_column=17
    sheet[get_column_letter(max_column) + str(summary_row)] = "max"
    sheet[get_column_letter(max_column+1) + str(summary_row)] = "strategy"
    sheet[get_column_letter(max_column + 2) + str(summary_row)] = "improved percentage"
    sheet[get_column_letter(max_column + 3) + str(summary_row)] = "value"
    for column_name in column_number_map:
        summary_one_column_max(sheet, total_row, column_number, column_number_map, row_number_map, max_column,summary_row,column_name)
        summary_row+=1



def summary_one_column_max(sheet,total_row,column_number,column_number_map,row_number_map,max_column,summary_row,column_name):
    comparison_column=column_number_map[column_name]
    improved_percentage_row = total_row + 2
    current_column_number = comparison_column
    max_percentage = -1
    max_value=-1
    max_percentage_cell = [None, None,None]
    while sheet[get_column_letter(current_column_number) + str(improved_percentage_row)].value is not None:
        current_cell_percentage = percentage_to_float(sheet[get_column_letter(current_column_number) + str(improved_percentage_row)].value)
        current_cell_value = sheet[get_column_letter(current_column_number) + str(improved_percentage_row-2)].value
        if current_cell_percentage > max_percentage:
            max_percentage = current_cell_percentage
            max_value=current_cell_value
            max_percentage_cell[0] = str(improved_percentage_row)
            max_percentage_cell[1] = get_column_letter(current_column_number)
            max_percentage_cell[2] = get_column_letter(current_column_number- (current_column_number- 4) % column_number) #compute corresponding strategy column

        current_column_number += column_number
    sheet[max_percentage_cell[1] + max_percentage_cell[0]].fill = green_fill
    max_strategy = sheet[max_percentage_cell[2] + str(row_number_map["strategy"])].value
    max_percentage_value=sheet[max_percentage_cell[1] + max_percentage_cell[0]].value
    sheet[get_column_letter(max_column) + str(summary_row + 1)].value = column_name
    sheet[get_column_letter(max_column+1) + str(summary_row + 1)].value = max_strategy
    sheet[get_column_letter(max_column + 2) + str(summary_row + 1)].value = max_percentage_value
    sheet[get_column_letter(max_column + 3) + str(summary_row + 1)].value = max_value


def compute_improved_percentage_for_common_solving_time(sheet,oirginal_st_column_number,row,column_number):
    current_st_column_number=oirginal_st_column_number
    while sheet[get_column_letter(current_st_column_number) + str(row)].value is not None:
        target_st_value = float(sheet[get_column_letter(current_st_column_number) + str(row)].value) # avg_t_cs
        oirginal_st_value = float(sheet[get_column_letter(current_st_column_number + 1) + str(row)].value)  # avg_t_ocs
        improve_percentage = (oirginal_st_value - target_st_value) / oirginal_st_value
        sheet[get_column_letter(current_st_column_number) + str(row + 2)].value = float_to_percentage(improve_percentage)
        # output a row combine numerical data and improved percentage for latex
        sheet[get_column_letter(current_st_column_number) + str(row + 4)].value = (sheet[get_column_letter(current_st_column_number) + str(row)].value +"\n"+ "(" +float_to_percentage(
                                                                                       improve_percentage)) + ")"
        current_st_column_number = current_st_column_number + column_number
def compute_improved_percentage_solving_time(sheet,oirginal_st_column_number,row,column_number):
    oirginal_st_value = float(sheet[get_column_letter(oirginal_st_column_number) + str(row)].value)
    current_st_column_number = oirginal_st_column_number + column_number
    while sheet[get_column_letter(current_st_column_number) + str(row)].value is not None:
        target_st_value = float(sheet[get_column_letter(current_st_column_number) + str(row)].value)
        improve_percentage = (oirginal_st_value - target_st_value) / oirginal_st_value
        sheet[get_column_letter(current_st_column_number) + str(row + 2)].value = float_to_percentage(improve_percentage)
        # output a row combine numerical data and improved percentage for latex
        sheet[get_column_letter(current_st_column_number) + str(row + 4)].value = (sheet[get_column_letter(current_st_column_number) + str(row)].value + "\n"+ "("+
                                                                                   float_to_percentage(improve_percentage)) + ")"
        current_st_column_number = current_st_column_number + column_number
def compute_improved_percentage(sheet,oirginal_column_number,row,column_number):
    oirginal_st_value = float(sheet[get_column_letter(oirginal_column_number) + str(row)].value)
    current_st_column_number = oirginal_column_number + column_number
    while sheet[get_column_letter(current_st_column_number) + str(row)].value is not None:
        target_st_value = float(sheet[get_column_letter(current_st_column_number) + str(row)].value)
        improve_percentage = (target_st_value-oirginal_st_value) / oirginal_st_value
        sheet[get_column_letter(current_st_column_number) + str(row + 2)].value = float_to_percentage(
            improve_percentage)

        #output a row combine numerical data and improved percentage for latex
        sheet[get_column_letter(current_st_column_number) + str(row + 4)].value = str(sheet[get_column_letter(current_st_column_number) + str(row)].value)+"\n"+"("+float_to_percentage(improve_percentage)+")"

        current_st_column_number = current_st_column_number + column_number

def get_column_letter(col_num):
    letter = ''
    while col_num:
        col_num, remainder = divmod(col_num - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter

def compute_average_solving_time(list_total, list_solved, list_time,list_common_time,common_solving_count,list_common_original_time,sat_solving_time_list,sat_solving_count_list,unsat_solving_time_list,unsat_solving_count_list):
    average_solving_time_list=[]
    average_solving_time_solved_list = []
    average_common_solving_time_list = []
    average_common_original_solving_time_list=[]
    average_safe_solving_time_list=[]
    average_unsafe_solving_time_list=[]
    #compute average solving time
    for x, y in zip(list_total[1:], list_time):
        if isinstance(x, str):
            average_solving_time_list.append(x)
        elif x == 0:
            average_solving_time_list.append(0)
        else:
            average_solving_time_list.append(y / x)
    #compute average solving time for safe
    for x, y, z in zip(list_total[1:], sat_solving_time_list, sat_solving_count_list):
        if isinstance(x, str):
            average_safe_solving_time_list.append(x)
        elif z == 0:
            average_safe_solving_time_list.append(0)
        else:
            average_safe_solving_time_list.append(y / z)
    #compute average solving time for unsafe
    for x, y, z in zip(list_total[1:], unsat_solving_time_list, unsat_solving_count_list):
        if isinstance(x, str):
            average_unsafe_solving_time_list.append(x)
        elif z == 0:
            average_unsafe_solving_time_list.append(0)
        else:
            average_unsafe_solving_time_list.append(y / z)
    #compute average solving time for solved
    for t,s,time in zip(list_total[1:],list_solved[1:], list_time):
        if isinstance(t, str):
            average_solving_time_solved_list.append(t)
        elif s==0:
            average_solving_time_solved_list.append(0)
        else:
            numerator = (time - (t - s) * benchmark_timeout)
            average_solving_time_solved_list.append(numerator/ s)
    #compute average solving time for common solved
    for x, y, z in zip(list_total[1:], list_common_time,common_solving_count):
        if isinstance(x, str):
            average_common_solving_time_list.append(x)
        elif z == 0:
            average_common_solving_time_list.append(0)
        else:
            average_common_solving_time_list.append(y / z)
    # compute average solving time for original common solved
    for x, y, z in zip(list_total[1:], list_common_original_time, common_solving_count):
        if isinstance(x, str):
            average_common_original_solving_time_list.append(x)
        elif z == 0:
            average_common_original_solving_time_list.append(0)
        else:
            average_common_original_solving_time_list.append(y / z)


    return average_solving_time_list,average_solving_time_solved_list,average_common_solving_time_list,average_common_original_solving_time_list,average_safe_solving_time_list,average_unsafe_solving_time_list
def draw_solving_time_scatter(excel_file, compare_benchmark_name):
    # Read the Excel file into a Pandas DataFrame
    solvability_dict = read_solvability_dict(excel_file)
    plot_folder = "/home/cheli243/PycharmProjects/HintsLearning/benchmarks/final-linear-evaluation/data_summary/scatter_plots"
    shutil.rmtree(plot_folder)
    scatter_folder = make_dirct(plot_folder)

    if "eldarica_symex_original_satisfiability" in solvability_dict.keys():
        apptainer_data_pair = ["eldarica_symex_original", "vb_eldarica_symex_prioritize"]
    elif "eldarica_CEGAR_original_satisfiability" in solvability_dict.keys():
        apptainer_data_pair = ["eldarica_CEGAR_original", "vb_eldarica_CEGAR_prioritize"]

    comparison_pairs = [["eldarica_abstract_off", "vb_eldarica_abstract_off_prioritizing_SEH"],
                        ["eldarica_abstract_off", "vb_eldarica_abstract_off_prioritizing_rank"],
                        ["eldarica_abstract_off", "vb_eldarica_abstract_off_pruning_rank"],
                        ["eldarica_abstract_off", "vb_eldarica_abstract_off_pruning_score"],
                        # ["eldarica_symex_original", "vb_eldarica_symex_prioritize"],
                        # ["eldarica_symex_original", "pf_eldarica_symex"]
                        # ["eldarica_CEGAR_original", "vb_eldarica_CEGAR_prioritize"],
                        ] + [apptainer_data_pair]

    axis_name_map = {"eldarica_abstract_off": "Fixed",
                     "vb_eldarica_abstract_off_prioritizing_SEH": "prioritizing_score+",
                     "vb_eldarica_abstract_off_prioritizing_rank": "Prioritizing_score",
                     "vb_eldarica_abstract_off_pruning_rank": "Pruning_rank",
                     "vb_eldarica_abstract_off_pruning_score": "Pruning_score",
                     "eldarica_symex_original": "Fixed_priority_function", "vb_eldarica_symex_prioritize": "MUS-guided_priority_function",
                     "eldarica_CEGAR_original": "Fixed_priority_function", "vb_eldarica_CEGAR_prioritize": "MUS-guided_priority_function"
                     }
    for pair in comparison_pairs:
        if "symex" in pair[0]:
            engine = "Symex"
        else:
            engine = "CEGAR"
        # draw common solvig time scatter
        original_solving_time_list_common = []
        strategy_solving_time_list_common = []
        satisfiability_list_common = []
        file_name_list_common = []
        # draw all solvig time scatter
        original_solving_time_list_all = []
        strategy_solving_time_list_all = []
        satisfiability_list_all = []
        file_name_list_all = []

        for name, original_s, strategy_s, original_st, strategy_st in zip(solvability_dict["file_name"],
                                                                          solvability_dict[pair[0] + "_satisfiability"],
                                                                          solvability_dict[pair[1] + "_satisfiability"],
                                                                          solvability_dict[pair[0] + "_solving_time"],
                                                                          solvability_dict[pair[1] + "_solving_time"]):
            vb_satisfiability = virtual_best_satisfiability_from_list([original_s, strategy_s])
            # file_name_list_all.append(name)
            # satisfiability_list_all.append(vb_satisfiability)
            # original_solving_time_list_all.append(original_st)
            # strategy_solving_time_list_all.append(strategy_st)
            # collect all solvable solving time and solvability
            if vb_satisfiability != "unknown":
                file_name_list_all.append(name)
                satisfiability_list_all.append(vb_satisfiability)
                original_solving_time_list_all.append(original_st)
                strategy_solving_time_list_all.append(strategy_st)
                # collect common solvable solving time and solvability
                if original_s == strategy_s:
                    file_name_list_common.append(name)
                    satisfiability_list_common.append(vb_satisfiability)
                    original_solving_time_list_common.append(original_st)
                    strategy_solving_time_list_common.append(strategy_st)

        engine = "CEGAR" if "CEGAR" in compare_benchmark_name else "Symex"

        if "union" in compare_benchmark_name:
            MUS_dataset= "union"
        elif "common" in compare_benchmark_name:
            MUS_dataset = "intersection"
        else:
            MUS_dataset = "single"

        if "REHPlus" in compare_benchmark_name:
            priority_function = "REHPlus"
        elif "REHMinus" in compare_benchmark_name:
            priority_function = "REHMinus"
        elif "SEHPlus" in compare_benchmark_name:
            priority_function = "SEHPlus"
        elif "SEHMinus" in compare_benchmark_name:
            priority_function = "SEHMinus"
        elif "score" in compare_benchmark_name:
            priority_function = "score"
        elif "rank" in compare_benchmark_name:
            priority_function = "rank"
        elif "twoQueue02" in compare_benchmark_name:
            priority_function = "twoQueue02"
        elif "twoQueue05" in compare_benchmark_name:
            priority_function = "twoQueue05"
        elif "twoQueue08" in compare_benchmark_name:
            priority_function = "twoQueue08"
        elif "schedule10" in compare_benchmark_name:
            priority_function = "schedule10"
        elif "schedule100" in compare_benchmark_name:
            priority_function = "schedule100"
        elif "schedule1000" in compare_benchmark_name:
            priority_function = "schedule1000"
        else:
            priority_function = "random"

        title_text="Algorithm:"+engine+", MUS dataset:"+MUS_dataset+", Priority function:"+priority_function

        all_solving_time_name="Solving time for all problems" + "<br>" + title_text
        common_solving_time_name="Solving time for commonly solved problems" + "<br>" + title_text
        save_file_name=scatter_plot(x_data=original_solving_time_list_common, y_data=strategy_solving_time_list_common,
                     z_data=satisfiability_list_common,
                     x_axis=axis_name_map[pair[0]], y_axis=axis_name_map[pair[1]], folder=scatter_folder,
                     data_text=file_name_list_common,
                     name=common_solving_time_name, scale="log",compare_benchmark_name=compare_benchmark_name)
        save_file_name=scatter_plot(x_data=original_solving_time_list_all, y_data=strategy_solving_time_list_all,
                     z_data=satisfiability_list_all,
                     x_axis=axis_name_map[pair[0]], y_axis=axis_name_map[pair[1]], folder=scatter_folder,
                     data_text=file_name_list_all,
                     name=all_solving_time_name, scale="log",compare_benchmark_name=compare_benchmark_name)
    return save_file_name


def read_solvability_dict(excel_file, sheet_name="data"):
    # Read the Excel file into a Pandas DataFrame
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0)

    # Convert the DataFrame to a dictionary
    solvability_dict = df.to_dict(orient='list')
    return solvability_dict


def read_json_file(file, json_obj):
    try:
        loaded_graph = json.load(file)
    except ValueError as e:
        file = file.name
        print("json.load() error", file)
        file_name = file[:file.find("smt2") + 4]
        copy_relative_files(file_name,
                            "/home/cheli243/PycharmProjects/HintsLearning/benchmarks/benchmark_statistics-debug")
        return json_obj
    for field in loaded_graph:
        # print(bcolors.GRENN + str(field) + str(loaded_graph[field]) + bcolors.RESET)
        json_obj[str(field)] = loaded_graph[field]
    return json_obj


def read_graph_generation_log(f, json_obj):
    for l in f.readlines():
        for g in ["CDHG", "CG"]:
            if g in l:
                json_obj.update({g + "_time_consumption": int(l[l.find(":") + 1:l.find("milliseconds")])})
    return json_obj


def read_smt2_category(f, json_obj):
    first_line = f.readline()[1:]
    json_obj["category"] = os.path.dirname(first_line)
    return json_obj


def read_files(file_list, file_type="solvability.JSON", read_function=read_json_file, disable_tqdm=False):
    '''notice: this is a generator and the returned interator can only be used once'''
    for file in tqdm(file_list, desc="read " + file_type, disable=disable_tqdm):
        file_name = file[:-len(".zip")]
        json_file = file_name + "." + file_type if len(file_type) != 0 else file_name
        try:
            unzip_file(json_file + ".zip")
        except:
            exception_folder = make_dirct(os.path.dirname(os.path.dirname(file)) + "/unzip_exceptions")
            copy_relative_files(file_name, exception_folder)
        json_obj = {}
        json_obj["file_name"] = json_file[:-len(file_type) - 1]
        if os.path.exists(json_file):
            json_obj["file_size"] = os.path.getsize(json_file)
            json_obj["file_size_h"] = convert_bytes(os.path.getsize(json_file))
            with open(json_file) as f:
                read_function(f, json_obj)
                # delete unziped file
            if os.path.exists(json_file + ".zip"):
                os.remove(json_file)
            yield json_obj
        else:
            yield json_obj


def get_sumary_folder(folder):
    summary_folder = os.path.dirname(folder) + "/" + os.path.basename(folder) + "_summary"
    make_dirct(summary_folder)
    return summary_folder


def copy_relative_files(file_name, folder):
    for f in glob.glob(file_name + "*"):
        copy(f, folder)


def get_min_max_solving_time(solving_time_dict, data_dict, object, func=min):
    solving_option, solving_time = select_key_with_value_condition(solving_time_dict, func)
    solving_time_cegar_interation_number = int(
        float(object[solving_option.replace("solvingTime", "cegarIterationNumber")][0]))
    solving_time_generated_predicate_number = int(
        float(object[solving_option.replace("solvingTime", "generatedPredicateNumber")][0]))
    solving_time_average_predicate_size = int(
        float(object[solving_option.replace("solvingTime", "averagePredicateSize")][0]))
    solving_time_predicate_generator_time = int(
        float(object[solving_option.replace("solvingTime", "predicateGeneratorTime")][0]))

    data_dict[func.__name__ + "_solving_time_option"].append(solving_option.replace("solvingTime_", ""))
    data_dict[func.__name__ + "_solving_time (s)"].append(solving_time / 1000)
    data_dict[func.__name__ + "_solving_time_cegar_interation_number"].append(
        solving_time_cegar_interation_number)
    data_dict[func.__name__ + "_solving_time_generated_predicate_number"].append(
        solving_time_generated_predicate_number)
    data_dict[func.__name__ + "_solving_time_average_predicate_size"].append(
        solving_time_average_predicate_size)
    data_dict[func.__name__ + "_solving_time_predicate_generator_time"].append(
        solving_time_predicate_generator_time)
    return solving_option, solving_time / 1000


def get_solving_time_dict(object):
    solving_time_dict = {}
    for field in object:
        if "solvingTime" in field:
            solving_time_dict[field] = int(float(object[field][0]))
            if solving_time_dict[field] == -1:
                solving_time_dict[field] = 10800000
    return solving_time_dict

def percentage_to_float(percentage):
    return float(percentage.strip('%')) / 100

def virtual_best_satisfiability_from_list(s):
    if "safe" in s and "unsafe" in s:
        return "unsound error"
    elif "safe" in s:
        return "safe"
    elif "unsafe" in s:
        return "unsafe"
    elif all(element == "miss info" for element in s):
        return "miss info"
    else:
        return "unknown"
